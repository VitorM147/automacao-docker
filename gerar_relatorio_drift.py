"""
Relatorio de Drift de Imagens Docker - Grupo DPSP
Le o checkpoint.json ou Excel existente e gera relatorio analitico
com metricas de padronizacao, versoes EOL, uso de :latest, etc.
"""

import json
import os
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference

CHECKPOINT_FILE = "checkpoint.json"
OUTPUT = "relatorio_drift_docker.xlsx"

# Imagens/versoes conhecidas como EOL
EOL_PATTERNS = [
    (r"node[:/]14", "Node.js 14 (EOL Abr/2023)"),
    (r"node[:/]12", "Node.js 12 (EOL Abr/2022)"),
    (r"node[:/]16", "Node.js 16 (EOL Set/2023)"),
    (r"python[:/]3\.7", "Python 3.7 (EOL Jun/2023)"),
    (r"python[:/]3\.6", "Python 3.6 (EOL Dez/2021)"),
    (r"python[:/]3\.8", "Python 3.8 (EOL Out/2024)"),
    (r"python[:/]2", "Python 2.x (EOL Jan/2020)"),
    (r"openjdk[:/]8", "OpenJDK 8 (EOL)"),
    (r"openjdk[:/]11", "OpenJDK 11 (LTS fim 2024)"),
    (r"java[:/]8", "Java 8 (EOL)"),
    (r"golang[:/]1\.1[0-8]\b", "Go < 1.19 (EOL)"),
    (r"ruby[:/]2\.", "Ruby 2.x (EOL)"),
    (r"php[:/]7\.", "PHP 7.x (EOL Nov/2022)"),
    (r"centos[:/]7", "CentOS 7 (EOL Jun/2024)"),
    (r"centos[:/]6", "CentOS 6 (EOL Nov/2020)"),
    (r"ubuntu[:/]18\.04", "Ubuntu 18.04 (EOL Mai/2023)"),
    (r"ubuntu[:/]16\.04", "Ubuntu 16.04 (EOL Abr/2021)"),
    (r"debian[:/]stretch", "Debian 9 Stretch (EOL)"),
    (r"debian[:/]buster", "Debian 10 Buster (EOL Jun/2024)"),
    (r"amazoncorretto[:/]8", "Amazon Corretto 8 (legado)"),
    (r"amazoncorretto[:/]11", "Amazon Corretto 11 (LTS fim 2027, considerar migrar)"),
]

# Mapeamento de runtime a partir do nome da imagem
RUNTIME_PATTERNS = [
    (r"node", "Node.js"),
    (r"python", "Python"),
    (r"openjdk|java|jdk|jre|corretto|temurin|eclipse-temurin", "Java"),
    (r"maven|gradle", "Java (Build)"),
    (r"golang|go:", "Go"),
    (r"ruby", "Ruby"),
    (r"php", "PHP"),
    (r"dotnet|aspnet", ".NET"),
    (r"nginx", "Nginx"),
    (r"httpd|apache", "Apache"),
    (r"postgres", "PostgreSQL"),
    (r"mysql|mariadb", "MySQL/MariaDB"),
    (r"mongo", "MongoDB"),
    (r"redis", "Redis"),
    (r"docker", "Docker"),
    (r"terraform", "Terraform"),
    (r"alpine", "Alpine (base)"),
    (r"ubuntu", "Ubuntu (base)"),
    (r"debian", "Debian (base)"),
    (r"centos|rocky|alma", "RHEL-based"),
    (r"tomcat", "Tomcat"),
]


def load_data():
    """Tenta carregar do checkpoint.json, senao le do Excel final."""
    # Tenta checkpoint primeiro
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        results = data.get("results", [])
        print(f"Fonte: {CHECKPOINT_FILE}")
        print(f"Registros carregados: {len(results)}")
        return results

    # Senao, le do Excel
    excel_file = "dockerfiles_dpsp.xlsx"
    if not os.path.exists(excel_file):
        print(f"ERRO: Nem {CHECKPOINT_FILE} nem {excel_file} encontrados.")
        return []

    from openpyxl import load_workbook
    wb = load_workbook(excel_file)
    ws = wb.active
    results = []
    headers = [c.value for c in ws[1]]
    col_map = {
        "Projeto": "project", "Namespace": "namespace", "Branch": "branch",
        "Ambiente": "environment", "Origem": "origin", "Arquivo": "file",
        "Imagem Docker": "image", "Sistema Operacional": "os", "URL": "url",
    }
    for row in ws.iter_rows(min_row=2, values_only=False):
        record = {}
        for i, cell in enumerate(row):
            header = headers[i] if i < len(headers) else None
            key = col_map.get(header)
            if key:
                record[key] = cell.value or ""
        if record.get("image"):
            results.append(record)

    print(f"Fonte: {excel_file}")
    print(f"Registros carregados: {len(results)}")
    return results


def detect_runtime(image):
    img = image.lower()
    for pattern, runtime in RUNTIME_PATTERNS:
        if re.search(pattern, img):
            return runtime
    return "Outro"


def check_eol(image):
    img = image.lower()
    for pattern, desc in EOL_PATTERNS:
        if re.search(pattern, img):
            return desc
    return None


def uses_latest(image):
    return image.endswith(":latest") or ":" not in image


def analyze(results):
    """Gera todas as metricas a partir dos resultados."""
    # Filtrar apenas registros com imagem real
    valid = [r for r in results if r["image"] != "(nenhum FROM)"]

    # --- Metricas gerais ---
    total_images = len(valid)
    unique_images = set(r["image"] for r in valid)
    projects_with_docker = set(r["project"] for r in valid)

    # --- Por ambiente ---
    by_env = Counter(r["environment"] for r in valid)

    # --- Runtime ---
    runtimes = Counter(detect_runtime(r["image"]) for r in valid)

    # --- Top imagens ---
    image_counter = Counter(r["image"] for r in valid)
    top_images = image_counter.most_common(20)

    # --- :latest ---
    latest_records = [r for r in valid if uses_latest(r["image"])]
    latest_by_env = Counter(r["environment"] for r in latest_records)

    # --- EOL ---
    eol_records = []
    for r in valid:
        eol = check_eol(r["image"])
        if eol:
            eol_records.append({**r, "eol_reason": eol})
    eol_by_reason = Counter(r["eol_reason"] for r in eol_records)

    # --- SO ---
    os_counter = Counter(r["os"] for r in valid)

    # --- Padronizacao ---
    top3_count = sum(c for _, c in image_counter.most_common(3))
    padronizacao = (top3_count / total_images * 100) if total_images > 0 else 0

    # --- Prod especifico ---
    prod_records = [r for r in valid if r["environment"] == "Prod"]
    prod_images = Counter(r["image"] for r in prod_records)
    prod_latest = [r for r in prod_records if uses_latest(r["image"])]
    prod_eol = [r for r in prod_records if check_eol(r["image"])]

    return {
        "total_images": total_images,
        "unique_images": len(unique_images),
        "projects_with_docker": len(projects_with_docker),
        "by_env": by_env,
        "runtimes": runtimes,
        "top_images": top_images,
        "latest_records": latest_records,
        "latest_by_env": latest_by_env,
        "latest_count": len(latest_records),
        "eol_records": eol_records,
        "eol_by_reason": eol_by_reason,
        "os_counter": os_counter,
        "padronizacao": padronizacao,
        "prod_latest": prod_latest,
        "prod_eol": prod_eol,
        "prod_top": prod_images.most_common(10),
    }


# ============================================================
# HELPERS DE ESTILO
# ============================================================
HFONT = Font(bold=True, color="FFFFFF", size=11)
HFILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BRD = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def write_header(ws, cols, widths=None):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HFONT
        c.fill = HFILL
        c.alignment = HALIGN
        c.border = BRD
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w


def write_row(ws, row, vals, fills=None):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = BRD
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if fills and col in fills:
            c.fill = fills[col]


# ============================================================
# GERACAO DAS ABAS
# ============================================================
def create_resumo(wb, m):
    ws = wb.active
    ws.title = "Resumo Executivo"
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25

    data = [
        ("INVENTARIO DOCKER - GRUPO DPSP", ""),
        ("", ""),
        ("Projetos com Docker", m["projects_with_docker"]),
        ("Total de registros (imagem+branch)", m["total_images"]),
        ("Imagens unicas", m["unique_images"]),
        ("", ""),
        ("DISTRIBUICAO POR AMBIENTE", ""),
        ("Producao (Prod)", m["by_env"].get("Prod", 0)),
        ("QA / Homologacao", m["by_env"].get("QA", 0)),
        ("Desenvolvimento (Dev)", m["by_env"].get("Dev", 0)),
        ("", ""),
        ("INDICADORES DE RISCO", ""),
        ("Imagens com :latest (anti-pattern)", m["latest_count"]),
        ("  -> em Producao", len(m["prod_latest"])),
        ("Imagens com versao EOL", len(m["eol_records"])),
        ("  -> em Producao", len(m["prod_eol"])),
        ("", ""),
        ("PADRONIZACAO", ""),
        ("Indice (top 3 imagens cobrem X%)", f"{m['padronizacao']:.1f}%"),
    ]

    for i, (label, val) in enumerate(data, 1):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=val)
        c1.border = BRD
        c2.border = BRD
        if label and label.isupper():
            c1.font = Font(bold=True, size=12)
        if "RISCO" in str(label):
            c1.font = Font(bold=True, size=12, color="CC0000")
        if ":latest" in str(label) or "EOL" in str(label):
            if isinstance(val, int) and val > 0:
                c2.fill = RED_FILL


def create_runtime(wb, m):
    ws = wb.create_sheet("Por Runtime")
    write_header(ws, ["Runtime", "Quantidade", "% do Total"], [30, 15, 15])

    total = m["total_images"]
    for i, (runtime, count) in enumerate(m["runtimes"].most_common(), 2):
        pct = (count / total * 100) if total > 0 else 0
        write_row(ws, i, [runtime, count, f"{pct:.1f}%"])


def create_top_images(wb, m):
    ws = wb.create_sheet("Top 20 Imagens")
    write_header(ws, ["Imagem", "Quantidade", "SO Detectado", "Usa :latest", "EOL"],
                 [50, 15, 25, 12, 35])

    for i, (image, count) in enumerate(m["top_images"], 2):
        eol = check_eol(image) or ""
        latest = "SIM" if uses_latest(image) else ""
        from gerar_relatorio_drift import detect_runtime
        fills = {}
        if latest == "SIM":
            fills[4] = YELLOW_FILL
        if eol:
            fills[5] = RED_FILL
        # Detectar SO inline
        os_name = "Nao identificado"
        img_lower = image.lower()
        if "alpine" in img_lower:
            os_name = "Alpine"
        elif "slim" in img_lower:
            os_name = "Debian (slim)"
        elif "debian" in img_lower or "buster" in img_lower or "bullseye" in img_lower:
            os_name = "Debian"
        elif "ubuntu" in img_lower or "jammy" in img_lower:
            os_name = "Ubuntu"
        write_row(ws, i, [image, count, os_name, latest, eol], fills)


def create_latest(wb, m):
    ws = wb.create_sheet("Uso de latest")
    write_header(ws, ["Projeto", "Namespace", "Branch", "Ambiente", "Imagem", "URL"],
                 [30, 30, 20, 10, 40, 60])

    for i, r in enumerate(m["latest_records"], 2):
        env_fill = {}
        if r["environment"] == "Prod":
            env_fill[4] = RED_FILL
        write_row(ws, i, [
            r["project"], r["namespace"], r["branch"],
            r["environment"], r["image"], r["url"]
        ], env_fill)
        # Hyperlink na URL
        ws.cell(row=i, column=6).hyperlink = r["url"]
        ws.cell(row=i, column=6).font = Font(color="0563C1", underline="single")


def create_eol(wb, m):
    ws = wb.create_sheet("Versoes EOL")
    write_header(ws, ["Projeto", "Namespace", "Branch", "Ambiente", "Imagem", "Motivo EOL", "URL"],
                 [30, 30, 20, 10, 40, 35, 60])

    for i, r in enumerate(m["eol_records"], 2):
        env_fill = {}
        if r["environment"] == "Prod":
            env_fill[4] = RED_FILL
        env_fill[6] = RED_FILL
        write_row(ws, i, [
            r["project"], r["namespace"], r["branch"],
            r["environment"], r["image"], r["eol_reason"], r["url"]
        ], env_fill)
        ws.cell(row=i, column=7).hyperlink = r["url"]
        ws.cell(row=i, column=7).font = Font(color="0563C1", underline="single")


def create_so(wb, m):
    ws = wb.create_sheet("Por SO")
    write_header(ws, ["Sistema Operacional", "Quantidade", "% do Total"], [35, 15, 15])

    total = m["total_images"]
    for i, (so, count) in enumerate(m["os_counter"].most_common(), 2):
        pct = (count / total * 100) if total > 0 else 0
        write_row(ws, i, [so, count, f"{pct:.1f}%"])


def create_prod(wb, m):
    ws = wb.create_sheet("Producao - Detalhes")
    write_header(ws, ["Imagem", "Quantidade", ":latest", "EOL"], [50, 15, 12, 35])

    for i, (image, count) in enumerate(m["prod_top"], 2):
        eol = check_eol(image) or ""
        latest = "SIM" if uses_latest(image) else ""
        fills = {}
        if latest == "SIM":
            fills[3] = YELLOW_FILL
        if eol:
            fills[4] = RED_FILL
        write_row(ws, i, [image, count, latest, eol], fills)


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 50)
    print("  Relatorio de Drift - Docker DPSP")
    print("=" * 50)

    results = load_data()
    if not results:
        print("Nenhum dado encontrado.")
        return

    print("Analisando dados...")
    m = analyze(results)

    print(f"\n  Projetos com Docker: {m['projects_with_docker']}")
    print(f"  Imagens unicas: {m['unique_images']}")
    print(f"  Uso de :latest: {m['latest_count']} ({len(m['prod_latest'])} em Prod)")
    print(f"  Versoes EOL: {len(m['eol_records'])} ({len(m['prod_eol'])} em Prod)")
    print(f"  Padronizacao (top 3): {m['padronizacao']:.1f}%")

    print(f"\nGerando Excel...")
    wb = Workbook()

    create_resumo(wb, m)
    create_runtime(wb, m)
    create_top_images(wb, m)
    create_latest(wb, m)
    create_eol(wb, m)
    create_so(wb, m)
    create_prod(wb, m)

    # Filtros em todas as abas
    for ws in wb.worksheets:
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT)
    print(f"\n  Relatorio gerado: {OUTPUT}")
    print(f"  Abas: Resumo Executivo, Por Runtime, Top 20 Imagens,")
    print(f"        Uso de latest, Versoes EOL, Por SO, Producao - Detalhes")


if __name__ == "__main__":
    main()
