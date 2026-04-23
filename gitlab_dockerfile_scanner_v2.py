"""
Scanner de Dockerfiles do GitLab - Grupo DPSP (v2)
Versao com ThreadPool adaptativo + Rate Limiter + Checkpoint.
Continua de onde parou caso seja interrompido.
"""

import requests
import base64
import re
import sys
import json
import time
import threading
import os
from dotenv import load_dotenv

load_dotenv()
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CONFIGURACAO
# ============================================================
GITLAB_URL = "https://gitlab.com"
PRIVATE_TOKEN = os.environ.get("GITLAB_TOKEN", "SEU_TOKEN_AQUI")
GROUP_ID = "grupo-dpsp"
OUTPUT_FILE = "dockerfiles_dpsp.xlsx"
PARTIAL_FILE = "dockerfiles_dpsp_parcial.xlsx"
CHECKPOINT_FILE = "checkpoint.json"
MAX_WORKERS = 10
CHECKPOINT_INTERVAL = 10  # salva checkpoint a cada N projetos
EXCEL_INTERVAL = 50       # salva excel parcial a cada N projetos
# ============================================================

HEADERS = {"PRIVATE-TOKEN": PRIVATE_TOKEN}

OS_MAP = {
    "alpine": "Alpine Linux", "ubuntu": "Ubuntu", "debian": "Debian",
    "centos": "CentOS", "fedora": "Fedora", "amazonlinux": "Amazon Linux",
    "oraclelinux": "Oracle Linux", "rockylinux": "Rocky Linux",
    "almalinux": "AlmaLinux", "archlinux": "Arch Linux",
    "opensuse": "openSUSE", "photon": "VMware Photon OS",
    "busybox": "BusyBox", "scratch": "Scratch (sem OS)",
    "distroless": "Distroless (Debian-based)",
    "mcr.microsoft.com/windows": "Windows",
    "mcr.microsoft.com/dotnet": "Debian (.NET)",
    "node": "Debian (Node.js)", "python": "Debian (Python)",
    "golang": "Debian (Go)", "ruby": "Debian (Ruby)",
    "php": "Debian (PHP)", "openjdk": "Debian (OpenJDK)",
    "eclipse-temurin": "Debian (Temurin)",
    "maven": "Debian (Maven)", "gradle": "Debian (Gradle)",
    "nginx": "Debian (Nginx)", "httpd": "Debian (Apache)",
    "postgres": "Debian (PostgreSQL)", "mysql": "Debian (MySQL)",
    "mongo": "Debian (MongoDB)", "redis": "Debian (Redis)",
    "tomcat": "Debian (Tomcat)",
}


# ============================================================
# ADAPTIVE RATE LIMITER
# ============================================================
class AdaptiveRateLimiter:
    """
    Rate limiter adaptativo que le os headers do GitLab
    e ajusta velocidade + workers em tempo real.
    """

    def __init__(self):
        self._lock = threading.Lock()
        self._remaining = 300
        self._limit = 300
        self._reset_at = 0.0
        self._total_requests = 0
        self._total_waits = 0
        self._semaphore = threading.Semaphore(MAX_WORKERS)

    def _update_from_headers(self, headers):
        """Atualiza estado a partir dos headers da resposta."""
        remaining = headers.get("RateLimit-Remaining")
        limit = headers.get("RateLimit-Limit")
        reset_at = headers.get("RateLimit-Reset")
        if remaining is not None:
            self._remaining = int(remaining)
        if limit is not None:
            self._limit = int(limit)
        if reset_at is not None:
            self._reset_at = float(reset_at)

    def _get_delay(self):
        """Calcula delay baseado no remaining."""
        if self._remaining > 150:
            return 0
        if self._remaining > 80:
            return 0.05
        if self._remaining > 30:
            return 0.2
        # remaining < 30: espera ate o reset
        wait = max(0, self._reset_at - time.time()) + 1
        self._total_waits += 1
        print(f"    [RATE] Remaining={self._remaining}, aguardando {wait:.0f}s ate reset...")
        return wait

    def get_recommended_workers(self):
        """Retorna quantos workers devem estar ativos."""
        with self._lock:
            if self._remaining > 150:
                return MAX_WORKERS
            if self._remaining > 80:
                return max(5, MAX_WORKERS // 2)
            if self._remaining > 30:
                return 2
            return 1

    def get(self, url, params=None, timeout=30):
        """Faz GET respeitando o rate limit adaptativo."""
        self._semaphore.acquire()
        try:
            with self._lock:
                delay = self._get_delay()
            if delay > 0:
                time.sleep(delay)

            resp = requests.get(url, headers=HEADERS, params=params, timeout=timeout)

            with self._lock:
                self._update_from_headers(resp.headers)
                self._total_requests += 1

            # Se recebeu 429, respeita Retry-After
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 60))
                print(f"    [RATE] HTTP 429! Aguardando {retry_after}s (Retry-After)...")
                time.sleep(retry_after)
                resp = requests.get(url, headers=HEADERS, params=params, timeout=timeout)
                with self._lock:
                    self._update_from_headers(resp.headers)
                    self._total_requests += 1

            return resp
        finally:
            self._semaphore.release()

    def stats(self):
        with self._lock:
            return {
                "total_requests": self._total_requests,
                "remaining": self._remaining,
                "limit": self._limit,
                "total_waits": self._total_waits,
            }


# Instancia global
rate_limiter = AdaptiveRateLimiter()


# ============================================================
# CHECKPOINT MANAGER
# ============================================================
class CheckpointManager:
    """Salva e carrega progresso para retomar de onde parou."""

    def __init__(self, filepath=CHECKPOINT_FILE):
        self._filepath = filepath
        self._lock = threading.Lock()

    def load(self):
        """Retorna (set de project_ids processados, lista de results salvos)."""
        if not os.path.exists(self._filepath):
            return set(), []
        try:
            with open(self._filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            ids = set(data.get("processed_ids", []))
            results = data.get("results", [])
            print(f"  [CHECKPOINT] Carregado: {len(ids)} projetos ja processados, {len(results)} registros")
            return ids, results
        except Exception as e:
            print(f"  [CHECKPOINT] Erro ao carregar: {e}")
            return set(), []

    def save(self, processed_ids, results):
        """Salva estado atual."""
        with self._lock:
            data = {
                "processed_ids": list(processed_ids),
                "results": results,
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "total_projects": len(processed_ids),
                "total_results": len(results),
            }
            try:
                with open(self._filepath, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"  [CHECKPOINT] Erro ao salvar: {e}")

    def clear(self):
        if os.path.exists(self._filepath):
            os.remove(self._filepath)
            print("  [CHECKPOINT] Arquivo removido (execucao completa)")


# ============================================================
# FUNCOES DE DETECCAO (mesmas do v1)
# ============================================================
def detect_os(image_name):
    img = image_name.lower()
    if "-alpine" in img or ":alpine" in img:
        return "Alpine Linux"
    if "-slim" in img:
        return "Debian (slim)"
    if any(v in img for v in ("-buster", "-bullseye", "-bookworm")):
        return "Debian"
    if any(v in img for v in ("-jammy", "-focal", "-noble")):
        return "Ubuntu"
    for key, os_name in OS_MAP.items():
        if key in img:
            return os_name
    return "Nao identificado"


def classify_environment(branch_name):
    b = branch_name.lower().strip()
    prod_exact = {"main", "master", "production", "prod", "release"}
    if b in prod_exact or b.startswith("release/") or b.startswith("hotfix/"):
        return "Prod"
    qa_keywords = ("qa", "homolog", "staging", "stg", "hml", "uat", "test", "homol")
    if any(k in b for k in qa_keywords):
        return "QA"
    return "Dev"


def extract_from_instructions(content):
    froms = []
    for line in content.splitlines():
        line = line.strip()
        m = re.match(r"^FROM\s+(.+?)(\s+AS\s+.+)?$", line, re.IGNORECASE)
        if m:
            froms.append(m.group(1).strip())
    return froms


def extract_images_from_ci(ci_content):
    images = []
    for m in re.finditer(r"^\s*image:\s*['\"]?([^\s'\"#]+)", ci_content, re.MULTILINE):
        img = m.group(1).strip()
        if img and not img.startswith("$") and not img.startswith("{"):
            images.append(img)
    for m in re.finditer(r"^\s*-?\s*name:\s*['\"]?([^\s'\"#]+)", ci_content, re.MULTILINE):
        img = m.group(1).strip()
        if img and not img.startswith("$") and not img.startswith("{"):
            if ":" in img or "/" in img:
                images.append(img)
    for m in re.finditer(r"^\s*-\s+['\"]?([a-zA-Z0-9][^\s'\"#]*:[^\s'\"#]+)", ci_content, re.MULTILINE):
        img = m.group(1).strip()
        if not img.startswith("$"):
            images.append(img)
    seen = set()
    unique = []
    for img in images:
        if img not in seen:
            seen.add(img)
            unique.append(img)
    return unique


# ============================================================
# FUNCOES DE API (usando rate_limiter)
# ============================================================
def get_group_id(group_path):
    encoded = requests.utils.quote(group_path, safe="")
    resp = rate_limiter.get(GITLAB_URL + "/api/v4/groups/" + encoded)
    resp.raise_for_status()
    data = resp.json()
    print(f"  Grupo: {data.get('full_path', '')} (ID: {data['id']})")
    return data["id"]


def get_all_projects(group_id):
    if isinstance(group_id, str) and not group_id.isdigit():
        group_id = get_group_id(group_id)
    projects = []
    page = 1
    while True:
        resp = rate_limiter.get(
            GITLAB_URL + "/api/v4/groups/" + str(group_id) + "/projects",
            params={
                "per_page": 100, "page": page,
                "include_subgroups": "true",
                "archived": "false", "with_shared": "false",
            },
        )
        resp.raise_for_status()
        data = resp.json()
        if not data:
            break
        projects.extend(data)
        print(f"  Pagina {page}: +{len(data)} (total: {len(projects)})")
        page += 1
    return projects


def get_all_branches(project_id):
    branches = []
    page = 1
    while True:
        url = GITLAB_URL + "/api/v4/projects/" + str(project_id) + "/repository/branches"
        resp = rate_limiter.get(url, params={"per_page": 100, "page": page})
        if resp.status_code in (404, 403):
            break
        if resp.status_code != 200:
            break
        data = resp.json()
        if not data:
            break
        branches.extend([b["name"] for b in data])
        page += 1
    return branches


def find_dockerfiles(project_id, branch):
    dockerfiles = []
    page = 1
    while True:
        url = GITLAB_URL + "/api/v4/projects/" + str(project_id) + "/repository/tree"
        resp = rate_limiter.get(
            url,
            params={"ref": branch, "recursive": "true", "per_page": 100, "page": page},
        )
        if resp.status_code != 200:
            break
        data = resp.json()
        if not data or not isinstance(data, list):
            break
        for item in data:
            if item.get("type") == "blob" and "dockerfile" in item.get("name", "").lower():
                dockerfiles.append(item["path"])
        page += 1

    if not dockerfiles:
        url = GITLAB_URL + "/api/v4/projects/" + str(project_id) + "/repository/files/Dockerfile"
        resp = rate_limiter.get(url, params={"ref": branch})
        if resp.status_code == 200:
            dockerfiles.append("Dockerfile")

    return dockerfiles


def get_file_content(project_id, file_path, branch):
    encoded = requests.utils.quote(file_path, safe="")
    url = GITLAB_URL + "/api/v4/projects/" + str(project_id) + "/repository/files/" + encoded
    resp = rate_limiter.get(url, params={"ref": branch})
    resp.raise_for_status()
    content_b64 = resp.json().get("content", "")
    return base64.b64decode(content_b64).decode("utf-8", errors="replace")


# ============================================================
# PROCESSAMENTO POR PROJETO (executado em cada thread)
# ============================================================
def process_branch(pid, proj_url, name, ns, branch):
    """Processa uma branch e retorna lista de results."""
    results = []
    env = classify_environment(branch)

    # --- Dockerfiles ---
    dockerfiles = find_dockerfiles(pid, branch)
    for df_path in dockerfiles:
        try:
            content = get_file_content(pid, df_path, branch)
            from_images = extract_from_instructions(content)
            df_url = proj_url + "/-/blob/" + branch + "/" + df_path

            if not from_images:
                results.append({
                    "project": name, "namespace": ns,
                    "branch": branch, "environment": env,
                    "origin": "Dockerfile", "file": df_path,
                    "image": "(nenhum FROM)", "os": "Nao identificado",
                    "url": df_url,
                })
                continue

            for image in from_images:
                os_det = detect_os(image)
                results.append({
                    "project": name, "namespace": ns,
                    "branch": branch, "environment": env,
                    "origin": "Dockerfile", "file": df_path,
                    "image": image, "os": os_det,
                    "url": df_url,
                })
                print(f"    [{env}] {branch} | Dockerfile: {image} ({os_det})")
        except Exception as e:
            print(f"    [ERRO] {branch}/{df_path}: {e}")

    # --- .gitlab-ci.yml ---
    try:
        ci_content = get_file_content(pid, ".gitlab-ci.yml", branch)
        ci_images = extract_images_from_ci(ci_content)
        ci_url = proj_url + "/-/blob/" + branch + "/.gitlab-ci.yml"

        for image in ci_images:
            os_det = detect_os(image)
            results.append({
                "project": name, "namespace": ns,
                "branch": branch, "environment": env,
                "origin": "Pipeline (.gitlab-ci.yml)", "file": ".gitlab-ci.yml",
                "image": image, "os": os_det,
                "url": ci_url,
            })
            print(f"    [{env}] {branch} | Pipeline: {image} ({os_det})")
    except Exception:
        pass

    return results


def process_project(proj):
    """Processa um projeto inteiro (todas as branches). Retorna (project_id, results)."""
    pid = proj["id"]
    name = proj["name"]
    proj_url = proj["web_url"]
    ns = proj.get("namespace", {}).get("full_path", "")

    branches = get_all_branches(pid)
    if not branches:
        return pid, []

    all_results = []
    for branch in branches:
        branch_results = process_branch(pid, proj_url, name, ns, branch)
        all_results.extend(branch_results)

    return pid, all_results


# ============================================================
# GERACAO DO EXCEL (mesma do v1)
# ============================================================
def generate_excel(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dockerfiles DPSP"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    brd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    env_colors = {
        "Prod": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "QA":   PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "Dev":  PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }
    origin_colors = {
        "Dockerfile": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "Pipeline (.gitlab-ci.yml)": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    }

    cols = ["Projeto", "Namespace", "Branch", "Ambiente", "Origem",
            "Arquivo", "Imagem Docker", "Sistema Operacional", "URL"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = brd

    for row, item in enumerate(results, 2):
        vals = [item["project"], item["namespace"], item["branch"],
                item["environment"], item["origin"], item["file"],
                item["image"], item["os"], item["url"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = brd
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 4:
                c.fill = env_colors.get(item["environment"], PatternFill())
            if col == 5:
                c.fill = origin_colors.get(item["origin"], PatternFill())
            if col == 9:
                c.hyperlink = val
                c.font = Font(color="0563C1", underline="single")

    widths = [30, 25, 20, 10, 25, 25, 40, 28, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)
    print(f"\nArquivo gerado: {output_file}")


# ============================================================
# MAIN - ORQUESTRADOR COM THREADS + CHECKPOINT
# ============================================================
def main():
    print("=" * 60)
    print("  Scanner de Dockerfiles - Grupo DPSP (v2)")
    print("  Modo: ThreadPool adaptativo + Checkpoint")
    print("=" * 60)

    checkpoint = CheckpointManager()
    processed_ids, saved_results = checkpoint.load()

    print(f"\nBuscando projetos do grupo '{GROUP_ID}'...")
    all_projects = get_all_projects(GROUP_ID)
    total = len(all_projects)
    print(f"  -> {total} projetos encontrados.")

    # Filtra projetos ja processados
    pending = [p for p in all_projects if p["id"] not in processed_ids]
    print(f"  -> {len(processed_ids)} ja processados (checkpoint)")
    print(f"  -> {len(pending)} pendentes\n")

    results = list(saved_results)  # copia dos resultados ja salvos
    proj_with_docker = 0
    skip_count = 0
    completed = len(processed_ids)
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {}
        for proj in pending:
            future = executor.submit(process_project, proj)
            futures[future] = proj

        for future in as_completed(futures):
            proj = futures[future]
            pid = proj["id"]
            name = proj["name"]
            ns = proj.get("namespace", {}).get("full_path", "")
            completed += 1

            try:
                project_id, project_results = future.result()
                processed_ids.add(project_id)

                if project_results:
                    results.extend(project_results)
                    proj_with_docker += 1
                    print(f"[{completed}/{total}] {ns}/{name} -> {len(project_results)} registros")
                else:
                    branches = get_all_branches(pid)
                    if not branches:
                        skip_count += 1
                        print(f"[{completed}/{total}] {ns}/{name} -> Sem branches")
                    else:
                        print(f"[{completed}/{total}] {ns}/{name} -> Nenhum Docker encontrado")

            except Exception as e:
                processed_ids.add(pid)
                print(f"[{completed}/{total}] {ns}/{name} -> ERRO: {e}")

            # Checkpoint a cada N projetos
            if completed % CHECKPOINT_INTERVAL == 0:
                checkpoint.save(processed_ids, results)
                elapsed = time.time() - start_time
                stats = rate_limiter.stats()
                remaining_projs = total - completed
                rate = completed / max(elapsed, 1)
                eta = remaining_projs / max(rate, 0.01)
                print(f"  [CHECKPOINT] Salvo: {completed}/{total} projetos | "
                      f"{len(results)} registros | "
                      f"API remaining: {stats['remaining']}/{stats['limit']} | "
                      f"ETA: {eta/60:.1f} min")

            # Excel parcial a cada 50 projetos
            if completed % EXCEL_INTERVAL == 0 and results:
                generate_excel(results, PARTIAL_FILE)
                print(f"  [EXCEL PARCIAL] {PARTIAL_FILE} atualizado ({len(results)} registros)")

    # Checkpoint final
    checkpoint.save(processed_ids, results)

    elapsed = time.time() - start_time
    stats = rate_limiter.stats()

    print(f"\n{'=' * 60}")
    print(f"  CONCLUIDO em {elapsed/60:.1f} minutos")
    print(f"  {completed} projetos processados")
    print(f"  {proj_with_docker} projetos com Docker")
    print(f"  {skip_count} projetos vazios")
    print(f"  {len(results)} registros no total")
    print(f"  {stats['total_requests']} requests a API")
    print(f"  {stats['total_waits']} pausas por rate limit")
    print(f"{'=' * 60}")

    if results:
        generate_excel(results, OUTPUT_FILE)
        checkpoint.clear()
    else:
        print("\nNenhum Dockerfile ou imagem Docker encontrada.")


if __name__ == "__main__":
    main()
