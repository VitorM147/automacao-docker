"""
Le o checkpoint.json e gera Excel com os dados ja coletados.
Nao interfere no script que esta rodando.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CHECKPOINT_FILE = "checkpoint.json"
OUTPUT = "dockerfiles_dpsp_parcial.xlsx"

with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
    data = json.load(f)

results = data.get("results", [])
print(f"Registros encontrados: {len(results)}")
print(f"Projetos processados: {data.get('total_projects', '?')}")
print(f"Ultimo checkpoint: {data.get('timestamp', '?')}")

if not results:
    print("Nenhum registro para gerar Excel.")
    exit()

wb = Workbook()
ws = wb.active
ws.title = "Dockerfiles DPSP (parcial)"

hfont = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
brd = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
env_colors = {
    "Prod": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "QA":   PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "Dev":  PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}
origin_colors = {
    "Dockerfile": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "Pipeline (.gitlab-ci.yml)": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}

cols = ["Projeto", "Namespace", "Branch", "Ambiente", "Origem",
        "Arquivo", "Imagem Docker", "Sistema Operacional", "URL"]
for i, h in enumerate(cols, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = halign
    c.border = brd

for row, item in enumerate(results, 2):
    vals = [item["project"], item["namespace"], item["branch"],
            item["environment"], item["origin"], item["file"],
            item["image"], item["os"], item["url"]]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = brd
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if col == 4:
            c.fill = env_colors.get(item["environment"], PatternFill())
        if col == 5:
            c.fill = origin_colors.get(item["origin"], PatternFill())
        if col == 9:
            c.hyperlink = val
            c.font = Font(color="0563C1", underline="single")

widths = [30, 25, 20, 10, 25, 25, 40, 28, 70]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

ws.auto_filter.ref = ws.dimensions
wb.save(OUTPUT)
print(f"\nExcel gerado: {OUTPUT} ({len(results)} registros)")
